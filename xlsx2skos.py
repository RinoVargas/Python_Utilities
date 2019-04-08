import sys
import argparse
import openpyxl
from jinja2 import Environment, PackageLoader, select_autoescape

DESCRIPTION = '''
Genera un vocabulario SKOS a partir de un archivo XLSX.
El archivo XLSX debe seguir la plantilla ubicada en templates/plantilla.xlsx
'''

def parse_args(args):
	'''Hacer parsing de los argumentos recibidos por línea de comando.'''
	
	parser = argparse.ArgumentParser(description=DESCRIPTION)
	parser.add_argument("source",
						metavar="ORIGEN",
						help="Nombre del archivo entrada")
	parser.add_argument("-t",
						"--tab",
						metavar="PESTAÑA",
						help="Nombre de la pestaña a leer en el archivo de entrada. Por defecto se lee la primera pestaña.")
	parser.add_argument("target",
						metavar="DESTINO",
						help="Nombre del archivo XML de salida")
	parser.add_argument("-f",
						"--template_file",
						help="Nombre del archivo de plantilla. Default: skos-xl.xml",
						default="skos-xl.xml",
						metavar="<filename>")
	parser.add_argument("-d",
						"--template_dir",
						help="Nombre de la carpeta de plantillas. Default: templates",
						default="templates",
						metavar="<dirname>")
#	parser.add_argument("--int_value",
#						help="display a square of a given number",
#						type=int)
#	parser.add_argument("--float_value",
#						help="display a square of a given number",
#						type=int)
#
#	parser.add_argument("-f",
#						"--flag",
#						help="Specify a flag",
#						action="store_true")
#						
#	parser.add_argument("--rating",
#						help="An option with a limited range of values",
#						choices=[1, 2, 3],
#						type=int)

	return parser.parse_args(args)
	
def load_data(args):
#
#	Cargar archivo .xlsx. 
#
	excel_document = openpyxl.load_workbook(args.source)

#
#	Si la no se ingresó el parametro --tab, se leerá desde la primera pestaña.
#
	if args.tab is None:
		first_cell_name = excel_document.sheetnames[0]
		excel_sheet = excel_document[first_cell_name]
	else:
		excel_sheet = excel_document[args.tab]
	
#
#	Apunta la fila 4 del Excel y extrae los valores de las columnas (A,B,C,D,E), Una vez extraído 
#	todas las columnas pasa a la siguiente fila.
#
#	Nota: Corta la iteración solo si la columna A (URI) no tiene valor (si es NoneType).
#

	row_num = 4
	concepts = list()
	metadata = excel_sheet.cell(1,2).value

	while (True):

	    if (excel_sheet.cell(row_num,1).value is None):
	        break
	    container = list()
	    for i in range(1,6):
	        container.append(excel_sheet.cell(row_num,i).value)
	    values = dict(	uri = container[0],
						definition_es = container[1], 
						prefLabel_es = container[2], 
						prefLabel_en = container[3], 
						broader = container[4])
	    concepts.append(values)            
	    row_num+=1

	return { "metadata": metadata, "concepts": concepts }


def render(args, data_dict):
	'''Escribir datos en archivo de salida usando plantilla.'''
	env = Environment(
		loader=PackageLoader('xlsx2skos', args.template_dir),
		autoescape=select_autoescape(['html', 'xml'])
	)
	template = env.get_template(args.template_file)
	rendered_text = template.render(**data_dict).encode("utf-8")
	with open(args.target, 'wb') as f:
		f.write(rendered_text)
	return


def main(args = sys.argv[1:]):
	args = parse_args(args)
	data_dict = load_data(args)
	render(args, data_dict)	

if __name__ == '__main__':
    main()